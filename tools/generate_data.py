import json
from pathlib import Path
from datetime import datetime

import openpyxl


BASE_DIR = Path(__file__).resolve().parent.parent
SOURCE_FILE = BASE_DIR / "data-source" / "中国常见食品营养参考_v3.xlsx"
OUTPUT_DIR = BASE_DIR / "dist" / "assets" / "data"


def is_non_empty_row(row):
    return any(v is not None and str(v).strip() != "" for v in row)


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return round(value, 2)
    return value


def build_records(sheet, header_row=1, data_start_row=2):
    headers = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
    records = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        if not is_non_empty_row(row):
            continue
        record = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            value = normalize_value(row[idx]) if idx < len(row) else None
            record[str(header).strip()] = value
        records.append(record)
    return records


def build_reference_rows(sheet):
    rows = []
    for row in sheet.iter_rows(values_only=True):
        if not is_non_empty_row(row):
            continue
        normalized = [normalize_value(v) for v in row]
        while normalized and normalized[-1] is None:
            normalized.pop()
        rows.append(normalized)
    return rows


def fill_portion_energy(records):
    for item in records:
        per100_kj = item.get("每100g热量(kJ)")
        per100_kcal = item.get("每100g能量(kcal)")
        portion_g = item.get("常见每份(g)")
        portion_kj = item.get("每份热量(kJ)")
        portion_kcal = item.get("每份能量(kcal)")
        if portion_g and isinstance(portion_g, (int, float)):
            if portion_kj is None and isinstance(per100_kj, (int, float)):
                item["每份热量(kJ)"] = int(round(per100_kj * portion_g / 100))
            if portion_kcal is None and isinstance(per100_kcal, (int, float)):
                item["每份能量(kcal)"] = int(round(per100_kcal * portion_g / 100))


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    main_sheet = wb["中国常见食品营养参考"]
    ingredient_sheet = wb["食材营养成分"]
    category_note_sheet = wb["分类说明"]
    usage_sheet = wb["使用说明"]
    daily_ref_sheet = wb["每日营养需求参考"]
    sport_ref_sheet = wb["运动能量消耗参考"]
    benchmark_sheet = wb["菜肴标杆对照表"]

    main_records = build_records(main_sheet)
    fill_portion_energy(main_records)

    ingredient_records = build_records(ingredient_sheet)
    fill_portion_energy(ingredient_records)

    payload = {
        "meta": {
            "source": str(SOURCE_FILE),
            "main_table_name": "中国常见食品营养参考",
            "main_count": len(main_records),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        },
        "main": main_records,
        "ingredients": ingredient_records,
        "notes": {
            "分类说明": build_records(category_note_sheet),
            "使用说明": build_records(usage_sheet),
        },
        "references": {
            "每日营养需求参考": build_reference_rows(daily_ref_sheet),
            "运动能量消耗参考": build_reference_rows(sport_ref_sheet),
            "菜肴标杆对照表": build_reference_rows(benchmark_sheet),
        },
    }

    json_path = OUTPUT_DIR / "nutrition-data.json"
    js_path = OUTPUT_DIR / "nutrition-data.js"

    with json_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    with js_path.open("w", encoding="utf-8") as f:
        f.write("window.NUTRITION_DATA = ")
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write(";")

    print("已生成:", json_path)
    print("已生成:", js_path)


if __name__ == "__main__":
    main()
